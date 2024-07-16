import json
import pandas as pd
import openpyxl as xl
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def load_json(json_path):
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return data

def json_to_dataframe(data):
    severity_data = data['severity']
    issues_data = data['issues']
    issues_txt = data['log']
    
    df_severity = pd.DataFrame([severity_data])
    df_issues = pd.DataFrame(issues_data)

    return df_severity, df_issues, issues_txt

def auto_fit_column_size(worksheet, columns=None, margin=2):
    for i, column_cells in enumerate(worksheet.columns):
        is_ok = False
        if columns is None:
            is_ok = True
        elif isinstance(columns, list) and i in columns:
            is_ok = True
            
        if is_ok:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = length + margin

    return worksheet

def merge_cells_by_newlines(sheet, start_row, text, alignment):
    start_col = 2
    end_col = 16

    for t in text:
        num_newlines = t.count('\n')
        num_lines = num_newlines + 1
        end_row = start_row + num_lines - 1
        start_col_letter = get_column_letter(start_col)
        end_col_letter = get_column_letter(end_col)
        merge_range = f'{start_col_letter}{start_row}:{end_col_letter}{end_row}'    
        sheet.merge_cells(merge_range)

        sheet[f'{start_col_letter}{start_row}'].value = t
        sheet[f'{start_col_letter}{start_row}'].alignment = alignment
        sheet[f'{start_col_letter}{start_row}'].border = Border(top=Side(style=None))
        sheet[f'{start_col_letter}{start_row}'].border = Border(bottom=Side(style=None))
        start_row = end_row + 1




def create_excel_report(df_severity, df_deps, issues_txt):
    today_date = datetime.today().strftime('%Y-%m-%d')
    workbook = xl.Workbook()
    sheet = workbook.active
    sheet.title = 'Snyk Report'

    insert_title = {
        "B": ['NO'],
        "C": ['플랫폼 구분'],
        "D": ['CWE Name'],
        "E": ['점검결과'],
        "F": ['진단일자'],
        "G": ['이행점검결과'],
        "H": ['Severity Level'],
        "I": ['Package'],
        "J": ['현재버전'],
        "K": ['패치버전'],
        "L": ['비고'],
        "M": ['조치계획'],
        "N": ['조치내역'],
        "O": ['조치일자'],
        "P": ['조치 Commit NO']
    }

    df_to_insert_title = pd.DataFrame(insert_title)

    insert_deps = {
        'B': df_deps.index + 1,
        'C': df_deps['platform'],
        'D': df_deps['cwe_name'],
        'E': ['취약'] * len(df_deps),
        'F': [today_date] * len(df_deps),
        'G': '',
        'H': df_deps['severity'].str.capitalize(),
        'I': df_deps['package'],
        'J': df_deps['now_ver'],
        'K': df_deps['to_ver'],
        'L': [f'외 {count}개' if int(count) > 0 else '' for count in df_deps['many']],
        'M':'',
        'N':'',
        'O':'',
        'P':''
    }
    df_to_insert_deps = pd.DataFrame(insert_deps)
    df_to_insert = pd.concat([df_to_insert_title, df_to_insert_deps], ignore_index=True)

    default_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    row_index = 1
    for i, row in df_to_insert.iterrows():
        row_index += 1
        for col, value in row.items():
            cell = sheet[f'{col}{row_index}']
            cell.value = value
            cell.alignment = center_alignment

    auto_fit_column_size(sheet)

    row_index += 5


    merge_cells_by_newlines(sheet, row_index, issues_txt, default_alignment)

    xlsx_name = ''
    xlsx_path = 'test_report.xlsx'
    workbook.save(xlsx_path)
    return xlsx_path

def json_to_xlsx(data):
    #data = load_json(json_path)
    df_severity, df_deps, issues_txt = json_to_dataframe(data)
    xlsx_path = create_excel_report(df_severity, df_deps, issues_txt)

